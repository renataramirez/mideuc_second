import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

folder = os.path.join("resultados", "06")

json_estructura = os.path.join(folder, "a1_estructura_soft.json")
json_intro = os.path.join(folder, "a1_introduccion_soft.json")
json_conc = os.path.join(folder, "a1_conclusion_cierre_soft.json")
json_prog = os.path.join(folder, "a1_progresion_textual_soft.json")

with open(json_estructura, "r", encoding="utf-8") as f:
    estructura = json.load(f)

with open(json_intro, "r", encoding="utf-8") as f:
    intro = json.load(f)

with open(json_conc, "r", encoding="utf-8") as f:
    conc = json.load(f)

with open(json_prog, "r", encoding="utf-8") as f:
    prog = json.load(f)

filas = []
contador_filas = 0  # Contador para filas en el DataFrame final

for item in estructura:
    id_texto = item["id_texto"]
    
    intro_est = item["resultado"].get("introduccion", {})
    conc_est = item["resultado"].get("conclusion_cierre", {})
    prog_est = item["resultado"].get("progresion_textual", {})
    
    intro_est_nivel = intro_est.get("puntaje", None)
    intro_est_just = intro_est.get("justificacion", "")
    
    conc_est_nivel = conc_est.get("puntaje", None)
    conc_est_just = conc_est.get("justificacion", "")
    
    prog_est_nivel = prog_est.get("puntaje", None)
    prog_est_just = prog_est.get("justificacion", "")
    
    intro_item = next((x for x in intro if x["id_texto"] == id_texto), None)
    conc_item = next((x for x in conc if x["id_texto"] == id_texto), None)
    prog_item = next((x for x in prog if x["id_texto"] == id_texto), None)
    
    intro_ind = intro_item["resultado"].get("introduccion", {}) if intro_item else {}
    conc_ind = conc_item["resultado"].get("conclusion_cierre", {}) if conc_item else {}
    prog_ind = prog_item["resultado"].get("progresion_textual", {}) if prog_item else {}
    
    intro_ind_nivel = intro_ind.get("puntaje", None)
    intro_ind_just = intro_ind.get("justificacion", "")
    
    conc_ind_nivel = conc_ind.get("puntaje", None)
    conc_ind_just = conc_ind.get("justificacion", "")
    
    prog_ind_nivel = prog_ind.get("puntaje", None)
    prog_ind_just = prog_ind.get("justificacion", "")
    
    # Filas para estructura (sin color)
    filas.append({
        "ID Texto": id_texto,
        "Introducción": intro_est_nivel,
        "Conclusión": conc_est_nivel,
        "Progresión Textual": prog_est_nivel
    })
    contador_filas += 1
    
    # Filas para justificación de estructura (sin color)
    filas.append({
        "ID Texto": "",
        "Introducción": intro_est_just,
        "Conclusión": conc_est_just,
        "Progresión Textual": prog_est_just
    })
    contador_filas += 1
    
    # Filas para consultas individuales - NIVEL (CON COLOR)
    filas.append({
        "ID Texto": "CONSULTA INDIVIDUAL",
        "Introducción": intro_ind_nivel,
        "Conclusión": conc_ind_nivel,
        "Progresión Textual": prog_ind_nivel
    })
    contador_filas += 1
    
    # Filas para consultas individuales - JUSTIFICACIÓN (CON COLOR)
    filas.append({
        "ID Texto": "",
        "Introducción": intro_ind_just,
        "Conclusión": conc_ind_just,
        "Progresión Textual": prog_ind_just
    })
    contador_filas += 1

df = pd.DataFrame(filas)

output_path = os.path.join(folder, "comparacion_puntajes_soft_final.xlsx")
df.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

# Ajustar dimensiones de columnas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40

# APLICAR COLORES A LAS FILAS DE CONSULTAS INDIVIDUALES
# Las filas 3, 7, 11, 15... (cada 4 filas empezando desde la 3) son las de consultas individuales
for fila in range(4, ws.max_row + 1, 4):
    for col in range(1, 5):  # Columnas A, B, C, D
        celda = ws.cell(row=fila, column=col)
        celda.fill = COLORES['introduccion']  # Puedes cambiar el color según necesites

# También colorear las filas de justificación de consultas individuales (siguiente fila)
for fila in range(5, ws.max_row + 1, 4):
    for col in range(1, 5):
        celda = ws.cell(row=fila, column=col)
        celda.fill = COLORES['introduccion']

# Aplicar alineación y wrap text a todas las celdas
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(output_path)

print(f"Excel generado en: {output_path}")
print("Comparación de puntajes completada.")