import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

folder = os.path.join("resultados", "06")

json_estructura = os.path.join(folder, "a1_estructura.json")
json_intro = os.path.join(folder, "a1_introduccion.json")
json_conc = os.path.join(folder, "a1_conclusion_cierre.json")
json_prog = os.path.join(folder, "a1_progresion_textual.json")

with open(json_estructura, "r", encoding="utf-8") as f:
    estructura = json.load(f)

with open(json_intro, "r", encoding="utf-8") as f:
    intro = json.load(f)

with open(json_conc, "r", encoding="utf-8") as f:
    conc = json.load(f)

with open(json_prog, "r", encoding="utf-8") as f:
    prog = json.load(f)

filas = []

for item in estructura:
    id_texto = item["id_texto"]
    
    intro_est = item["resultado"].get("introduccion", {})
    conc_est = item["resultado"].get("conclusion_cierre", {})
    prog_est = item["resultado"].get("progresion_textual", {})
    
    intro_est_nivel = intro_est.get("nivel", None)
    intro_est_just = intro_est.get("justificacion", "")
    
    conc_est_nivel = conc_est.get("nivel", None)
    conc_est_just = conc_est.get("justificacion", "")
    
    prog_est_nivel = prog_est.get("nivel", None)
    prog_est_just = prog_est.get("justificacion", "")
    
    intro_item = next((x for x in intro if x["id_texto"] == id_texto), None)
    conc_item = next((x for x in conc if x["id_texto"] == id_texto), None)
    prog_item = next((x for x in prog if x["id_texto"] == id_texto), None)
    
    intro_ind = intro_item["resultado"].get("introduccion", {}) if intro_item else {}
    conc_ind = conc_item["resultado"].get("conclusion_cierre", {}) if conc_item else {}
    prog_ind = prog_item["resultado"].get("progresion_textual", {}) if prog_item else {}
    
    intro_ind_nivel = intro_ind.get("nivel", None)
    intro_ind_just = intro_ind.get("justificacion", "")
    
    conc_ind_nivel = conc_ind.get("nivel", None)
    conc_ind_just = conc_ind.get("justificacion", "")
    
    prog_ind_nivel = prog_ind.get("nivel", None)
    prog_ind_just = prog_ind.get("justificacion", "")
    
    filas.append({
        "ID Texto": id_texto,
        "Introducción": intro_est_nivel,
        "Conclusión": conc_est_nivel,
        "Progresión Textual": prog_est_nivel
    })
    
    filas.append({
        "ID Texto": "",
        "Introducción": intro_est_just,
        "Conclusión": conc_est_just,
        "Progresión Textual": prog_est_just
    })
    
    filas.append({
        "ID Texto": "",
        "Introducción": intro_ind_nivel,
        "Conclusión": conc_ind_nivel,
        "Progresión Textual": prog_ind_nivel
    })
    
    filas.append({
        "ID Texto": "",
        "Introducción": intro_ind_just,
        "Conclusión": conc_ind_just,
        "Progresión Textual": prog_ind_just
    })

df = pd.DataFrame(filas)

output_path = os.path.join(folder, "comparacion_puntajes.xlsx")
df.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(output_path)

print(f"Excel generado en: {output_path}")
print("Comparación de puntajes completada.")